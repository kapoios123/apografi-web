"""
Απογραφή — Streamlit web εργαλείο (αντικαθιστά το tkinter main_gui.py)

Τύποι αποθηκών:
  * bins  -> MainActivity (locations/bins/μετρήσεις)
  * cases -> BoxWarehouseActivity (μοναδικός αριθμός, κατάσταση ΟΧΙ/Βρέθηκε)
  * retalia (cases) -> RetaliaActivity. ΔΟΜΗ: products/Ρετάλια/products/<ΤΟΠΟΘΕΣΙΑ>/<σειριακό>
        - σειριακό ΚΑΘΑΡΟ (π.χ. S3023), τοποθεσία ως γονέας (μοναδικότητα ανά τοποθεσία)
        - "διάσταση" = αρχική, "νέα_διάσταση" = διορθωμένη, "νέο"=true για app-added
"""

import io
import json
import datetime
import re

import pandas as pd
import streamlit as st
import firebase_admin
from firebase_admin import credentials, db
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

pd.set_option("future.no_silent_downcasting", True)

# ============================================================
# CONFIG
# ============================================================
WAREHOUSES = {
    "accessories":      "bins",
    "Bottero":          "bins",
    "Αναλώσιμα":        "bins",
    "systems":          "bins",
    "emergency":        "bins",
    "Φύλλα_Κιβωτίων":   "bins",
    "Αποθήκη_Κιβωτίων": "cases",
    "Ρετάλια":          "cases",   # cases-type με ειδικό import/export (δομή ανά τοποθεσία)
}

RETALIA_WAREHOUSES = {"Ρετάλια"}
LOCATIONS = ["MAIN", "GAR", "ATH"]


def sanitize_key(key) -> str:
    if not isinstance(key, str):
        key = str(key)
    key = key.strip()
    if key == "":
        return "empty_key"
    key = re.sub(r"[.$#\[\]/]", "_", key)
    key = re.sub(r"\s+", "_", key)
    return key


def admin_password():
    try:
        return st.secrets["admin"]["password"]
    except Exception:
        return None


def _clean_df(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(axis=1, how="all")
    df = df.dropna(axis=0, how="all")
    return df


def read_excel_clean(file) -> pd.DataFrame:
    """Διαβάζει Excel αγνοώντας κενές πρώτες γραμμές/στήλες (βρίσκει μόνη της τις επικεφαλίδες)."""
    raw = pd.read_excel(file, header=None)
    raw = raw.dropna(axis=1, how="all")
    raw = raw.dropna(axis=0, how="all")
    raw = raw.reset_index(drop=True)
    if raw.empty:
        return raw
    header = raw.iloc[0]
    df = raw.iloc[1:].copy()
    df.columns = [str(c).strip() for c in header]
    return df.reset_index(drop=True)


# ============================================================
# FIREBASE
# ============================================================
@st.cache_resource
def get_db_root():
    if not firebase_admin._apps:
        cred = credentials.Certificate(dict(st.secrets["firebase_admin"]))
        firebase_admin.initialize_app(
            cred, {"databaseURL": st.secrets["firebase"]["database_url"]}
        )
    return db.reference("/")


def upload_warehouse(warehouse: str, products_dict: dict):
    get_db_root().child("products").child(warehouse).set({"products": products_dict})


def read_warehouse(warehouse: str) -> dict:
    data = get_db_root().child("products").child(warehouse).child("products").get()
    return data or {}
    
def backup_warehouse_bytes(warehouse: str) -> bytes:
    """Ακριβές JSON αντίγραφο της τρέχουσας αποθήκης (για επαναφορά 1:1)."""
    data = read_warehouse(warehouse)
    payload = {
        "warehouse": warehouse,
        "backup_at": datetime.datetime.now().isoformat(),
        "products": data,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def restore_warehouse(warehouse: str, products_dict: dict):
    """Επαναφορά αποθήκης από backup JSON."""
    upload_warehouse(warehouse, products_dict)

def warehouse_kind(name: str) -> str:
    if name in RETALIA_WAREHOUSES:
        return "retalia"
    return WAREHOUSES.get(name, "bins")


def publish_warehouses_config() -> dict:
    cfg = {name: {"kind": warehouse_kind(name)} for name in WAREHOUSES}
    get_db_root().child("config").child("warehouses").set(cfg)
    return cfg


# ============================================================
# CONVERTERS
# ============================================================
def convert_bins(df: pd.DataFrame, warehouse: str, user_name: str = "import") -> dict:
    df = _clean_df(df).fillna(0)
    for col in df.columns:
        df[col] = df[col].apply(lambda x: str(x).upper().strip() if isinstance(x, str) else x)

    result = {}
    for _, row in df.iterrows():
        raw_code = str(row.get("ΚΩΔΙΚΟΣ", "")).strip()
        if raw_code == "" or raw_code.lower() == "nan":
            continue

        code = sanitize_key(raw_code)
        location = sanitize_key(str(row.get("ΑΠΟΘΗΚΗ", warehouse)).strip())
        bin_name = sanitize_key(str(row.get("Bin", row.get("Bins", "default"))).strip())
        shelf = str(row.get("ΡΑΦΙ", "") or "").strip()
        measurement_uid = sanitize_key(user_name)
        timestamp = datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3].replace(".", ":")

        try:
            initial_val = float(row.get("ΠΟΣΟΤΗΤΑ", 0) or 0)
        except (ValueError, TypeError):
            initial_val = 0.0

        if code not in result:
            result[code] = {
                "ΠΕΡΙΓΡΑΦΗ": row.get("ΠΕΡΙΓΡΑΦΗ", ""),
                "has_difference": initial_val != 0,
                "locations": {},
            }
        elif initial_val != 0:
            result[code]["has_difference"] = True

        product = result[code]
        product["locations"].setdefault(location, {"bins": {}})
        bins = product["locations"][location]["bins"]
        if bin_name not in bins:
            bins[bin_name] = {
                "αρχικό_απόθεμα": 0.0, "συνολική_μέτρηση": 0.0, "ραφι": "", "μετρήσεις": {},
            }

        if shelf:
            bins[bin_name]["ραφι"] = shelf
        bins[bin_name]["αρχικό_απόθεμα"] += initial_val
        bins[bin_name]["μετρήσεις"][measurement_uid] = {
            "user": user_name.upper(), "ποσότητα": 0, "timestamp": timestamp, "σχόλιο": "", "ραφι": "",
        }
    return result


def convert_cases(df: pd.DataFrame, warehouse: str) -> dict:
    df = _clean_df(df).fillna("")
    products = {}
    for _, row in df.iterrows():
        case_serial = sanitize_key(str(row.get("Case Serial Number", row.get("case", ""))))
        if case_serial in ("", "empty_key"):
            continue
        products[case_serial] = {
            "ΚΩΔΙΚΟΣ": str(row.get("ΚΩΔΙΚΟΣ", "")).strip(),
            "ΠΕΡΙΓΡΑΦΗ": str(row.get("ΠΕΡΙΓΡΑΦΗ", "")),
            "locations": {},
            "διάσταση": str(row.get("ΔΙΑΣΤΑΣΕΙΣ", "")),
            "τοποθεσια": str(row.get("ΑΠΟΘΗΚΗ", "")),
            "φύλλα": "", "Θεση": "", "κατάσταση": "ΟΧΙ", "μετρήσεις": {},
        }
    return products


def convert_retalia(df: pd.DataFrame, warehouse: str) -> dict:
    """
    Excel ρεταλιών -> ΝΕΑ δομή: { ΤΟΠΟΘΕΣΙΑ: { σειριακό: {...} } }
        'GAR:3023' -> τοποθεσία GAR, σειριακό S3023 (καθαρό)
    Διάσταση από ΜΟΝ.ΜΕΤΡ.
    """
    df = _clean_df(df).fillna("")
    products: dict = {}
    for _, row in df.iterrows():
        apothiki = str(row.get("ΑΠΟΘΗΚΗ", "")).strip()
        if ":" not in apothiki:
            continue  # χωρίς τοποθεσία δεν μπορούμε να φωλιάσουμε σωστά
        loc, num = apothiki.split(":", 1)
        loc = sanitize_key(loc.strip().upper())
        num = num.strip()
        if num == "" or loc in ("", "empty_key"):
            continue

        serial = sanitize_key("S" + num)
        dimension = str(row.get("ΜΟΝ.ΜΕΤΡ", "") or "").strip()

        products.setdefault(loc, {})[serial] = {
            "ΚΩΔΙΚΟΣ": str(row.get("ΚΩΔΙΚΟΣ", "")).strip(),
            "ΠΕΡΙΓΡΑΦΗ": str(row.get("ΠΕΡΙΓΡΑΦΗ", "")),
            "locations": {},
            "διάσταση": dimension,
            "τοποθεσια": loc,
            "φύλλα": "", "Θεση": "", "κατάσταση": "ΟΧΙ", "μετρήσεις": {},
        }
    return products


def build_json(df: pd.DataFrame, warehouse: str) -> dict:
    if warehouse in RETALIA_WAREHOUSES:
        return convert_retalia(df, warehouse)
    if WAREHOUSES[warehouse] == "bins":
        return convert_bins(df, warehouse)
    return convert_cases(df, warehouse)


def record_count(warehouse: str, built: dict) -> int:
    """Για ρετάλια (φωλιασμένα) μετράμε τα σειριακά, όχι τις τοποθεσίες."""
    if warehouse in RETALIA_WAREHOUSES:
        return sum(len(v) for v in built.values() if isinstance(v, dict))
    return len(built)


def validate_import(df: pd.DataFrame, warehouse: str, built: dict) -> list:
    warnings = []
    cols = [str(c).strip() for c in df.columns]

    def missing(required):
        return [c for c in required if c not in cols]

    if warehouse in RETALIA_WAREHOUSES:
        m = missing(["ΑΠΟΘΗΚΗ", "ΜΟΝ.ΜΕΤΡ"])
        if m:
            warnings.append(f"Λείπουν στήλες για ρετάλια: {', '.join(m)}")
        serials = [s for loc_d in built.values() if isinstance(loc_d, dict) for s in loc_d.keys()]
        s_keys = [k for k in serials if str(k).upper().startswith("S")]
        if serials and len(s_keys) < len(serials) * 0.5:
            warnings.append("Τα περισσότερα σειριακά δεν ξεκινούν με 'S'. Σίγουρα αρχείο ρεταλιών;")
        if "ΑΠΟΘΗΚΗ" in cols and not any(":" in str(v) for v in df["ΑΠΟΘΗΚΗ"].head(20)):
            warnings.append("Η στήλη ΑΠΟΘΗΚΗ δεν έχει μορφή 'GAR:3023' — σίγουρα αρχείο ρεταλιών;")

    elif WAREHOUSES[warehouse] == "cases":
        if "Case Serial Number" not in cols and "case" not in cols:
            warnings.append("Δεν βρέθηκε στήλη 'Case Serial Number'. Σίγουρα αρχείο κιβωτίων;")
        bad = [k for k in built if str(k).startswith("empty_key") or str(k).strip() == ""]
        if built and len(bad) > len(built) * 0.3:
            warnings.append("Πολλά κενά/άκυρα serial κιβωτίων — μάλλον λάθος αρχείο.")
        if "ΑΠΟΘΗΚΗ" in cols and any(":" in str(v) for v in df["ΑΠΟΘΗΚΗ"].head(20)):
            warnings.append("Η στήλη ΑΠΟΘΗΚΗ έχει μορφή 'GAR:1234' — μήπως είναι αρχείο ρεταλιών;")
        m = missing(["ΚΩΔΙΚΟΣ"])
        if m:
            warnings.append(f"Λείπει στήλη: {', '.join(m)}")

    else:  # bins
        m = missing(["ΚΩΔΙΚΟΣ", "ΑΠΟΘΗΚΗ", "ΠΟΣΟΤΗΤΑ"])
        if m:
            warnings.append(f"Λείπουν βασικές στήλες για bins: {', '.join(m)}. Σίγουρα σωστή αποθήκη;")
        if "Case Serial Number" in cols:
            warnings.append("Το αρχείο έχει 'Case Serial Number' — μήπως είναι αρχείο κιβωτίων;")

    if record_count(warehouse, built) == 0:
        warnings.append("Δεν παρήχθη καμία εγγραφή. Το αρχείο μάλλον είναι λάθος ή κενό.")

    return warnings


# ============================================================
# EXPORTS
# ============================================================
BLUE = PatternFill(start_color="B7D4F0", end_color="B7D4F0", fill_type="solid")
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


def _style_headers(ws):
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _idx(ws, name):
    header = [c.value for c in ws[1]]
    return header.index(name) if name in header else None


def _color_detail(ws):
    i_init, i_total = _idx(ws, "ΑΡΧΙΚΟ_ΑΠΟΘΕΜΑ"), _idx(ws, "ΣΥΝΟΛΙΚΗ_ΜΕΤΡΗΣΗ")
    if i_init is None or i_total is None:
        return
    for r in ws.iter_rows(min_row=2):
        try:
            iv = round(float(r[i_init].value or 0), 2)
            tv = round(float(r[i_total].value or 0), 2)
        except (ValueError, TypeError):
            continue
        r[i_total].fill = BLUE if tv == iv else (GREEN if tv > iv else RED)


def _color_diff_cell(ws):
    i_diff = _idx(ws, "ΔΙΑΦΟΡΑ")
    if i_diff is None:
        return
    for r in ws.iter_rows(min_row=2):
        try:
            d = round(float(r[i_diff].value or 0), 2)
        except (ValueError, TypeError):
            continue
        r[i_diff].fill = BLUE if d == 0 else (GREEN if d > 0 else RED)


def _highlight_diff_rows(ws):
    i_diff = _idx(ws, "ΔΙΑΦΟΡΑ")
    if i_diff is None:
        return
    for r in ws.iter_rows(min_row=2):
        try:
            d = float(r[i_diff].value or 0)
        except (ValueError, TypeError):
            d = 0.0
        if abs(d) > 1e-9:
            for c in r:
                c.fill = YELLOW


def _color_cases(ws):
    i_status, i_leaves = _idx(ws, "κατάσταση"), _idx(ws, "φύλλα")
    for r in ws.iter_rows(min_row=2):
        status = r[i_status].value if i_status is not None else ""
        leaves = r[i_leaves].value if i_leaves is not None else ""
        fill = None
        if leaves not in (None, "", "null"):
            fill = GREEN
        elif status == "Βρέθηκε":
            fill = BLUE
        elif status == "ΟΧΙ":
            fill = RED
        if fill:
            for c in r:
                c.fill = fill


def _color_retalia(ws):
    i_new = _idx(ws, "ΝΕΟ")
    i_status = _idx(ws, "ΚΑΤΑΣΤΑΣΗ")
    i_corr = _idx(ws, "ΔΙΟΡΘΩΘΗΚΕ")
    for r in ws.iter_rows(min_row=2):
        has_correction = i_corr is not None and str(r[i_corr].value or "").strip() == "ΝΑΙ"
        is_new = i_new is not None and str(r[i_new].value or "").strip() == "ΝΑΙ"
        status = (r[i_status].value if i_status is not None else "") or ""

        if has_correction:
            row_fill = YELLOW
        elif is_new:
            row_fill = GREEN
        elif status == "Βρέθηκε":
            row_fill = BLUE
        elif status == "ΟΧΙ":
            row_fill = RED
        else:
            row_fill = None

        if row_fill:
            for c in r:
                c.fill = row_fill


def _single_sheet(df: pd.DataFrame, sheet_name: str, color_fn=None) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    _style_headers(ws)
    if color_fn:
        color_fn(ws)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def _bins_detail_df(data: dict, only_differences: bool = False) -> pd.DataFrame:
    rows = []
    for code, pdata in data.items():
        if not isinstance(pdata, dict):
            continue
        if only_differences and pdata.get("has_difference") is not True:
            continue
        description = pdata.get("ΠΕΡΙΓΡΑΦΗ", "")
        for location, loc_data in (pdata.get("locations") or {}).items():
            for bin_name, bin_data in (loc_data.get("bins") or {}).items():
                if not isinstance(bin_data, dict):
                    continue
                initial_qty = float(bin_data.get("αρχικό_απόθεμα", 0) or 0)
                total_qty = float(bin_data.get("συνολική_μέτρηση", 0) or 0)
                if only_differences and initial_qty == total_qty:
                    continue

                new_shelves = []
                for _, m in (bin_data.get("μετρήσεις") or {}).items():
                    if isinstance(m, dict):
                        s = str(m.get("ραφι", "") or "").strip()
                        if s and s not in new_shelves:
                            new_shelves.append(s)

                row = {
                    "ΚΩΔΙΚΟΣ": code,
                    "ΠΕΡΙΓΡΑΦΗ": description,
                    "ΑΠΟΘΗΚΗ": location,
                    "BIN": bin_name,
                    "ΡΑΦΙ_ΠΑΛΙΟ": bin_data.get("ραφι", ""),
                    "ΡΑΦΙ_ΝΕΟ": ", ".join(new_shelves),
                    "ΑΡΧΙΚΟ_ΑΠΟΘΕΜΑ": initial_qty,
                    "ΣΥΝΟΛΙΚΗ_ΜΕΤΡΗΣΗ": total_qty,
                }
                for i, (_, m) in enumerate((bin_data.get("μετρήσεις") or {}).items(), start=1):
                    if isinstance(m, dict):
                        row[f"ΜΕΤΡΗΣΗ_{i}"] = (
                            f"{m.get('user','')} | {m.get('timestamp','')} | "
                            f"Ποσό: {m.get('ποσότητα',0)} | Σχόλιο: {m.get('σχόλιο','')} | "
                            f"Ραφι: {m.get('ραφι','')}"
                        )
                rows.append(row)
    return pd.DataFrame(rows)


def export_bins_differences(warehouse: str) -> bytes:
    df = _bins_detail_df(read_warehouse(warehouse), only_differences=True)
    if df.empty:
        return b""
    return _single_sheet(df, "Διαφορές", _color_detail)


def export_bins_full(warehouse: str) -> bytes:
    detail = _bins_detail_df(read_warehouse(warehouse), only_differences=False)
    if detail.empty:
        return b""

    by_loc = (
        detail.groupby(["ΚΩΔΙΚΟΣ", "ΠΕΡΙΓΡΑΦΗ", "ΑΠΟΘΗΚΗ"], as_index=False)[
            ["ΑΡΧΙΚΟ_ΑΠΟΘΕΜΑ", "ΣΥΝΟΛΙΚΗ_ΜΕΤΡΗΣΗ"]
        ].sum()
    )
    by_loc["ΔΙΑΦΟΡΑ"] = by_loc["ΣΥΝΟΛΙΚΗ_ΜΕΤΡΗΣΗ"] - by_loc["ΑΡΧΙΚΟ_ΑΠΟΘΕΜΑ"]

    src = detail[~detail["ΚΩΔΙΚΟΣ"].astype(str).str.contains("_C_", na=False)]
    summary = (
        src.groupby(["ΚΩΔΙΚΟΣ", "ΠΕΡΙΓΡΑΦΗ"], as_index=False)[
            ["ΑΡΧΙΚΟ_ΑΠΟΘΕΜΑ", "ΣΥΝΟΛΙΚΗ_ΜΕΤΡΗΣΗ"]
        ].sum()
    )
    summary["ΔΙΑΦΟΡΑ"] = summary["ΣΥΝΟΛΙΚΗ_ΜΕΤΡΗΣΗ"] - summary["ΑΡΧΙΚΟ_ΑΠΟΘΕΜΑ"]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="Αναλυτικά", index=False)
        by_loc.to_excel(writer, sheet_name="Ανά Τοποθεσία", index=False)
        summary.to_excel(writer, sheet_name="Σύνοψη", index=False)
    buf.seek(0)

    wb = load_workbook(buf)
    _style_headers(wb["Αναλυτικά"]); _color_detail(wb["Αναλυτικά"])
    _style_headers(wb["Ανά Τοποθεσία"]); _color_diff_cell(wb["Ανά Τοποθεσία"])
    _style_headers(wb["Σύνοψη"]); _highlight_diff_rows(wb["Σύνοψη"])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def export_cases(warehouse: str) -> bytes:
    data = read_warehouse(warehouse)
    rows = []
    for case_code, case_data in data.items():
        if not isinstance(case_data, dict):
            continue
        row = {
            "CASE": case_code,
            "ΚΩΔΙΚΟΣ": case_data.get("ΚΩΔΙΚΟΣ", ""),
            "ΠΕΡΙΓΡΑΦΗ": case_data.get("ΠΕΡΙΓΡΑΦΗ", ""),
            "διάσταση": case_data.get("διάσταση", ""),
            "τοποθεσια": case_data.get("τοποθεσια", ""),
            "κατάσταση": case_data.get("κατάσταση", ""),
            "φύλλα": case_data.get("φύλλα", ""),
        }
        measurements = case_data.get("μετρήσεις", {})
        if isinstance(measurements, dict):
            for i, (_, m) in enumerate(measurements.items(), start=1):
                if isinstance(m, dict):
                    row[f"ΜΕΤΡΗΣΗ_{i}"] = (
                        f"{m.get('user','')} | {m.get('timestamp','')} | "
                        f"Ποσό: {m.get('ποσότητα',0)} | Σχόλιο: {m.get('σχόλιο','')} | "
                        f"Θέση: {m.get('position', case_data.get('Θέση',''))}"
                    )
        rows.append(row)

    if not rows:
        return b""
    return _single_sheet(pd.DataFrame(rows), "Cases", _color_cases)


def export_retalia(warehouse: str) -> bytes:
    """Ρετάλια: φωλιασμένη δομή τοποθεσία -> σειριακό. Στήλες ΔΙΑΣΤΑΣΗ + ΝΕΑ_ΔΙΑΣΤΑΣΗ."""
    data = read_warehouse(warehouse)   # { ΤΟΠΟΘΕΣΙΑ: { σειριακό: {...} } }
    rows = []
    for loc, serials in data.items():
        if not isinstance(serials, dict):
            continue
        for serial, rdata in serials.items():
            if not isinstance(rdata, dict):
                continue
            row = {
                "SERIAL": serial,
                "ΤΟΠΟΘΕΣΙΑ": loc,
                "ΚΩΔΙΚΟΣ": rdata.get("ΚΩΔΙΚΟΣ", ""),
                "ΠΡΟΜΗΘΕΥΤΗΣ": rdata.get("προμηθευτής", ""),
                "ΔΙΑΣΤΑΣΗ": rdata.get("διάσταση", ""),
                "ΝΕΑ_ΔΙΑΣΤΑΣΗ": rdata.get("νέα_διάσταση", ""),
                "ΔΙΟΡΘΩΘΗΚΕ": "ΝΑΙ" if (str(rdata.get("νέα_διάσταση", "")).strip() != "" or rdata.get("διορθωμένο") is True) else "",
                "ΚΑΤΑΣΤΑΣΗ": rdata.get("κατάσταση", ""),
                "ΝΕΟ": "ΝΑΙ" if rdata.get("νέο") is True else "",
            }
            measurements = rdata.get("μετρήσεις", {})
            if isinstance(measurements, dict):
                for i, (_, m) in enumerate(measurements.items(), start=1):
                    if isinstance(m, dict):
                        row[f"ΜΕΤΡΗΣΗ_{i}"] = (
                            f"{m.get('user','')} | {m.get('timestamp','')} | "
                            f"Τοπ: {m.get('location','')}"
                        )
            rows.append(row)

    if not rows:
        return b""
    return _single_sheet(pd.DataFrame(rows), "Ρετάλια", _color_retalia)


# ============================================================
# UI
# ============================================================
def login_view():
    st.title("📦 Απογραφή — Είσοδος")
    if st.button("Είσοδος ως USER", use_container_width=True):
        st.session_state.role = "USER"
        st.rerun()
    st.divider()
    pwd = st.text_input("Κωδικός Admin", type="password")
    if st.button("Είσοδος ως ADMIN", use_container_width=True):
        if admin_password() and pwd == admin_password():
            st.session_state.role = "ADMIN"
            st.rerun()
        else:
            st.error("Λανθασμένος κωδικός.")


def app_view():
    role = st.session_state.role
    st.sidebar.success(f"Συνδεδεμένος ως **{role}**")
    if st.sidebar.button("Αποσύνδεση"):
        st.session_state.clear()
        st.rerun()

    checks_on = st.sidebar.toggle("Έλεγχοι import", value=True,
                                  help="Αν το κλείσεις, αγνοούνται οι προειδοποιήσεις και "
                                       "επιτρέπεται το upload χωρίς επιβεβαίωση.")

    if role == "ADMIN":
        st.sidebar.divider()
        st.sidebar.caption("Μετά από αλλαγή στη λίστα αποθηκών, δημοσίευσέ τη στην εφαρμογή:")
        if st.sidebar.button("📡 Δημοσίευση αποθηκών"):
            try:
                cfg = publish_warehouses_config()
                st.sidebar.success(f"Δημοσιεύτηκαν {len(cfg)} αποθήκες.")
            except Exception as e:
                st.sidebar.error(f"Σφάλμα: {e}")

    st.title("📦 Διαχείριση Απογραφής")
    warehouse = st.selectbox("Επέλεξε αποθήκη", list(WAREHOUSES.keys()))
    wtype = WAREHOUSES[warehouse]
    is_retalia = warehouse in RETALIA_WAREHOUSES
    st.caption(f"Τύπος αποθήκης: **{wtype}**" + ("  ·  ρετάλια" if is_retalia else ""))

    tabs = ["📊 Εξαγωγές"]
    if role == "ADMIN":
        tabs.append("🛠 Import / Upload")
    tab_objects = st.tabs(tabs)

    # ---------- ΕΞΑΓΩΓΕΣ ----------
    with tab_objects[0]:
        if wtype == "bins":
            c1, c2 = st.columns(2)
            with c1:
                if st.button("🔽 Μόνο Διαφορές", use_container_width=True):
                    _do_export(lambda: export_bins_differences(warehouse), f"DIFFERENCES_{warehouse}.xlsx")
            with c2:
                if st.button("📥 Πλήρης Εξαγωγή (3 φύλλα)", use_container_width=True):
                    _do_export(lambda: export_bins_full(warehouse), f"{warehouse}_export.xlsx")
            st.caption("Πλήρης = Αναλυτικά + Ανά Τοποθεσία + Σύνοψη (3 φύλλα).")
        else:  # cases / ρετάλια
            if st.button("🔽 Εξαγωγή Αποθήκης", use_container_width=True):
                fn = (lambda: export_retalia(warehouse)) if is_retalia else (lambda: export_cases(warehouse))
                _do_export(fn, f"{warehouse}_export.xlsx")

    # ---------- IMPORT / UPLOAD (admin) ----------
    if role == "ADMIN":
        with tab_objects[1]:
            st.write("1️⃣ Ανέβασε το Excel και χτίσε το JSON, 2️⃣ ανέβασέ το στη Firebase.")
            if is_retalia:
                st.info("Προσοχή: το upload ρεταλιών ΑΝΤΙΚΑΘΙΣΤΑ όλη την αποθήκη — "
                        "και όσα προστέθηκαν από την εφαρμογή. Χρήση κυρίως για αρχικό στήσιμο.")
            up = st.file_uploader("Excel αρχείο", type=["xlsx"], key=f"up_{warehouse}")
            if st.button("🔧 Χτίσιμο JSON", use_container_width=True):
                if not up:
                    st.warning("Διάλεξε πρώτα Excel.")
                else:
                    try:
                        df = read_excel_clean(up)
                        built = build_json(df, warehouse)
                        st.session_state.built_json = built
                        st.session_state.built_for = warehouse
                        st.session_state.built_warnings = validate_import(df, warehouse, built)
                        st.success(f"Διαβάστηκαν {record_count(warehouse, built)} εγγραφές για '{warehouse}'.")
                    except Exception as e:
                        st.error(f"Σφάλμα ανάγνωσης Excel: {e}")

            if st.session_state.get("built_for") == warehouse:
                built = st.session_state.get("built_json", {})
                warnings = st.session_state.get("built_warnings", []) if checks_on else []
                for w in warnings:
                    st.warning("⚠️ " + w)

                if not built and checks_on:
                    st.error("Δεν παρήχθη καμία εγγραφή — δες τις προειδοποιήσεις. "
                             "Μάλλον λάθος αρχείο ή αποθήκη. Το upload είναι κλειστό.")
                else:
                    st.download_button(
                        "💾 Κατέβασε το JSON (backup)",
                        data=json.dumps(built, ensure_ascii=False, indent=2).encode("utf-8"),
                        file_name=f"apothiki_{warehouse}.json",
                        mime="application/json",
                        use_container_width=True,
                    )

                    confirmed = True
                    if warnings:
                        confirmed = st.checkbox("Ναι, είμαι σίγουρος — ανέβασέ το παρά τις προειδοποιήσεις.")

                    # --- BACKUP πριν το upload (ΑΣΠΙΔΑ) ---
                    st.divider()
                    st.markdown("**🛟 Backup πριν το upload**")
                    if st.button("📥 Φτιάξε backup τρέχουσας αποθήκης", use_container_width=True):
                        try:
                            st.session_state.backup_bytes = backup_warehouse_bytes(warehouse)
                            st.session_state.backup_for = warehouse
                        except Exception as e:
                            st.error(f"Αποτυχία backup: {e}")

                    backup_ready = (st.session_state.get("backup_for") == warehouse
                                    and st.session_state.get("backup_bytes"))
                    if backup_ready:
                        st.download_button(
                            "💾 Κατέβασε το backup (JSON)",
                            data=st.session_state.backup_bytes,
                            file_name=f"BACKUP_{warehouse}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                            mime="application/json",
                            use_container_width=True,
                        )

                    backup_done = st.checkbox("✅ Έχω κατεβάσει το backup")

                    st.warning("Το Upload ΑΝΤΙΚΑΘΙΣΤΑ τα δεδομένα της αποθήκης στη Firebase.")
                    if st.button("📡 Upload στη Firebase", type="primary",
                                 use_container_width=True,
                                 disabled=not (confirmed and backup_done)):
                        try:
                            upload_warehouse(warehouse, built)
                            st.success(f"Ανέβηκαν στην αποθήκη '{warehouse}'.")
                        except Exception as e:
                            st.error(f"Αποτυχία upload: {e}")

            # --- ΕΠΑΝΑΦΟΡΑ από backup ---
            st.divider()
            st.markdown("**♻️ Επαναφορά από backup**")
            st.caption("Ανέβασε ένα backup JSON για να επαναφέρεις την αποθήκη όπως ήταν.")
            restore_file = st.file_uploader("Backup JSON", type=["json"], key=f"restore_{warehouse}")
            if st.button("♻️ Επαναφορά αποθήκης", use_container_width=True):
                if not restore_file:
                    st.warning("Διάλεξε πρώτα ένα backup JSON.")
                else:
                    try:
                        payload = json.load(restore_file)
                        products = payload.get("products", payload)
                        restore_warehouse(warehouse, products)
                        st.success(f"Επαναφέρθηκε η αποθήκη '{warehouse}'.")
                    except Exception as e:
                        st.error(f"Αποτυχία επαναφοράς: {e}")


def _do_export(fn, filename):
    try:
        data = fn()
    except Exception as e:
        st.error(f"Σφάλμα εξαγωγής: {e}")
        return
    if not data:
        st.info("Δεν βρέθηκαν δεδομένα για εξαγωγή.")
        return
    st.download_button(
        "⬇️ Κατέβασε το Excel",
        data=data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


def main():
    st.set_page_config(page_title="Απογραφή", page_icon="📦")
    if "role" not in st.session_state:
        login_view()
    else:
        app_view()


if __name__ == "__main__":
    main()
